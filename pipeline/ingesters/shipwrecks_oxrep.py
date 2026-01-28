"""
Oxford Roman Economy Project (OXREP) Shipwrecks ingester.

The OXREP shipwrecks database contains 2,000+ ancient Mediterranean
shipwrecks with cargo, dating, and location information.

Data source: https://oxrep.classics.ox.ac.uk/databases/shipwrecks_database/
Direct download: https://oxrep.web.ox.ac.uk/sitefiles/oxrep-shipwrecks-database-2.1-march-2017.xlsx
License: Academic (cite Strauss 2013)
API Key: Not required
"""

import json
import re
from pathlib import Path
from typing import Iterator, Optional, Dict, Any, List
from datetime import datetime
import time

from loguru import logger

from pipeline.ingesters.base import BaseIngester, ParsedSite, atomic_write_json
from pipeline.utils.http import fetch_with_retry, download_file


class OXREPShipwrecksIngester(BaseIngester):
    """
    Ingester for OXREP Shipwrecks Database.

    Downloads the Excel database directly from OXREP's file server.
    Also supports manual file placement for offline use.
    """

    source_id = "shipwrecks_oxrep"
    source_name = "OXREP Shipwrecks Database"

    # Direct Excel download (most reliable)
    EXCEL_URL = "https://oxrep.web.ox.ac.uk/sitefiles/oxrep-shipwrecks-database-2.1-march-2017.xlsx"

    # Fallback URLs
    BASE_URL = "https://oxrep.classics.ox.ac.uk"
    SEARCH_URL = f"{BASE_URL}/databases/shipwrecks_database/"

    # Manual file names to check (user can drop file here)
    MANUAL_FILES = [
        "oxrep-shipwrecks.xlsx",
        "oxrep-shipwrecks-database.xlsx",
        "oxrep-shipwrecks-database-2.1-march-2017.xlsx",
    ]

    def fetch(self) -> Path:
        """
        Fetch shipwreck data from OXREP.

        Tries in order:
        1. Check for manually placed Excel file in raw_data_dir
        2. Download Excel from OXREP file server
        3. Fall back to placeholder if download fails

        Returns:
            Path to JSON file with shipwreck data
        """
        dest_path = self.raw_data_dir / "shipwrecks_oxrep.json"
        excel_path = self.raw_data_dir / "oxrep-shipwrecks.xlsx"

        logger.info("Fetching OXREP Shipwrecks data...")
        self.report_progress(0, 3, "checking for local files...")

        all_wrecks = []

        # Step 1: Check for manually placed files
        for manual_file in self.MANUAL_FILES:
            manual_path = self.raw_data_dir / manual_file
            if manual_path.exists():
                logger.info(f"Found manual file: {manual_path}")
                excel_path = manual_path
                break

        # Step 2: Download if no manual file exists
        if not excel_path.exists():
            self.report_progress(1, 3, "downloading Excel file...")
            logger.info(f"Downloading OXREP Excel from {self.EXCEL_URL}")

            try:
                headers = {
                    "User-Agent": "AncientNerds/1.0 (Research Platform; academic research)",
                }
                download_file(self.EXCEL_URL, excel_path, force=True, decompress_gzip=False)
                logger.info(f"Downloaded to {excel_path}")
            except Exception as e:
                logger.warning(f"Failed to download Excel: {e}")

        # Step 3: Parse Excel file
        if excel_path.exists():
            self.report_progress(2, 3, "parsing Excel file...")
            all_wrecks = self._parse_excel(excel_path)

        # Create output
        if all_wrecks:
            logger.info(f"Parsed {len(all_wrecks):,} shipwrecks from Excel")
        else:
            logger.warning("Could not parse OXREP data. Creating placeholder.")
            all_wrecks = [{
                "id": "oxrep_placeholder",
                "name": "OXREP Database",
                "note": f"Download manually from {self.EXCEL_URL} and place in {self.raw_data_dir}",
                "estimated_records": 2000,
                "needs_manual_fetch": True,
            }]

        self.report_progress(3, 3, f"{len(all_wrecks):,} wrecks")

        # Save to JSON
        output = {
            "shipwrecks": all_wrecks,
            "metadata": {
                "source": "OXREP Shipwrecks Database",
                "source_url": "https://oxrep.classics.ox.ac.uk/databases/shipwrecks_database/",
                "download_url": self.EXCEL_URL,
                "fetched_at": datetime.utcnow().isoformat(),
                "total_records": len(all_wrecks),
                "data_type": "shipwrecks",
                "license": "Academic - cite Strauss 2013",
                "citation": "Strauss, J. (2013). Shipwrecks Database. Oxford Roman Economy Project.",
            }
        }

        atomic_write_json(dest_path, output)

        logger.info(f"Saved {len(all_wrecks):,} shipwrecks to {dest_path}")
        return dest_path

    def _parse_excel(self, excel_path: Path) -> List[Dict]:
        """Parse the OXREP Excel file."""
        try:
            import openpyxl
        except ImportError:
            logger.warning("openpyxl not installed. Run: pip install openpyxl")
            return []

        all_wrecks = []

        try:
            wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

            # Find the main data sheet
            sheet = None
            for sheet_name in wb.sheetnames:
                if 'wreck' in sheet_name.lower() or 'data' in sheet_name.lower():
                    sheet = wb[sheet_name]
                    break
            if sheet is None:
                sheet = wb.active

            logger.info(f"Parsing sheet: {sheet.title}")

            # Get headers from first row
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return []

            headers = [str(h).lower().strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]

            # Map common column names
            # Note: Order matters - check more specific patterns first
            col_map = {}
            for i, h in enumerate(headers):
                h_clean = h.replace(" ", "_").replace("-", "_")
                # Check for date columns FIRST (to avoid 'latest' matching 'lat')
                if any(x in h for x in ['earliest', 'date_start', 'date_from', 'min_date']):
                    col_map['date_start'] = i
                elif any(x in h for x in ['latest', 'date_end', 'date_to', 'max_date']):
                    col_map['date_end'] = i
                elif h in ['latitude', 'lat']:
                    col_map['lat'] = i
                elif h in ['longitude', 'lon', 'long', 'lng']:
                    col_map['lon'] = i
                elif any(x in h for x in ['locid', 'id', 'number', 'strauss']):
                    col_map['id'] = i
                elif any(x in h for x in ['wreck_name', 'wreck name']):
                    col_map['name'] = i
                elif any(x in h for x in ['site_name', 'site name', 'name']):
                    if 'name' not in col_map:  # Don't overwrite wreck_name
                        col_map['name'] = i
                elif any(x in h for x in ['region', 'area', 'location']):
                    col_map['region'] = i
                elif 'country' in h:
                    col_map['country'] = i
                elif 'depth' in h:
                    col_map['depth'] = i
                elif 'date' in h and 'date_start' not in col_map:
                    col_map['date'] = i
                elif any(x in h for x in ['cargo', 'contents', 'amphorae']):
                    col_map['cargo'] = i
                elif any(x in h for x in ['vessel', 'ship', 'type']):
                    col_map['vessel_type'] = i
                elif any(x in h for x in ['size', 'length', 'tonnage']):
                    col_map['size'] = i
                elif any(x in h for x in ['reference', 'bibliography', 'source']):
                    col_map['bibliography'] = i
                elif any(x in h for x in ['notes', 'comment', 'description']):
                    col_map['notes'] = i

            logger.info(f"Column mapping: {col_map}")

            # Parse data rows
            for row_idx, row in enumerate(rows[1:], start=2):
                if not any(row):  # Skip empty rows
                    continue

                def get_val(key, default=""):
                    if key in col_map and col_map[key] < len(row):
                        val = row[col_map[key]]
                        return val if val is not None else default
                    return default

                # Get ID
                wreck_id = str(get_val('id', '')).strip()
                if not wreck_id or wreck_id in ['', 'None']:
                    wreck_id = f"row_{row_idx}"
                wreck_id = f"oxrep_{wreck_id}".replace(" ", "_").replace("/", "_")

                # Get coordinates
                lat = self._parse_float(get_val('lat'))
                lon = self._parse_float(get_val('lon'))

                # Parse dates
                date_start = self._parse_year(get_val('date_start', get_val('date')))
                date_end = self._parse_year(get_val('date_end'))

                wreck = {
                    "id": wreck_id,
                    "name": str(get_val('name', f"Wreck {wreck_id}")),
                    "lat": lat,
                    "lon": lon,
                    "location_name": str(get_val('region', '')),
                    "country": str(get_val('country', '')),
                    "depth_meters": self._parse_float(get_val('depth')),
                    "date_sunk_start": date_start,
                    "date_sunk_end": date_end,
                    "vessel_type": str(get_val('vessel_type', '')),
                    "cargo_type": str(get_val('cargo', '')),
                    "size": str(get_val('size', '')),
                    "bibliography": str(get_val('bibliography', '')),
                    "notes": str(get_val('notes', '')),
                }

                all_wrecks.append(wreck)

            wb.close()

        except Exception as e:
            logger.error(f"Error parsing Excel file: {e}")
            import traceback
            traceback.print_exc()

        return all_wrecks

    def _extract_table_data(self, html: str) -> List[Dict]:
        """Extract data from HTML tables."""
        data = []

        # Find table rows
        row_pattern = re.compile(r'<tr[^>]*>(.*?)</tr>', re.DOTALL | re.IGNORECASE)
        cell_pattern = re.compile(r'<t[dh][^>]*>(.*?)</t[dh]>', re.DOTALL | re.IGNORECASE)

        rows = row_pattern.findall(html)
        headers = []

        for i, row in enumerate(rows):
            cells = cell_pattern.findall(row)
            # Clean HTML tags from cells
            cells = [re.sub(r'<[^>]+>', '', cell).strip() for cell in cells]

            if i == 0 and any(h.lower() in ['name', 'location', 'date', 'cargo', 'wreck'] for h in cells):
                headers = [h.lower().replace(' ', '_') for h in cells]
            elif headers and cells:
                row_data = dict(zip(headers, cells))
                if row_data:
                    data.append(row_data)

        return data

    def _parse_wreck_item(self, item: Dict) -> Optional[Dict]:
        """Parse a shipwreck item from various formats."""
        if not item or not isinstance(item, dict):
            return None

        # Try to get ID
        wreck_id = str(item.get("id", item.get("wreck_id", item.get("name", ""))))
        if not wreck_id:
            return None

        wreck_id = f"oxrep_{wreck_id}".replace(" ", "_").lower()

        # Get coordinates
        lat = item.get("lat", item.get("latitude", item.get("y")))
        lon = item.get("lon", item.get("lng", item.get("longitude", item.get("x"))))

        # Try to extract from geometry
        if lat is None and "geometry" in item:
            geom = item["geometry"]
            if geom.get("type") == "Point":
                coords = geom.get("coordinates", [])
                if len(coords) >= 2:
                    lon, lat = coords[0], coords[1]

        # Parse lat/lon
        try:
            lat = float(lat) if lat is not None else None
            lon = float(lon) if lon is not None else None
        except (ValueError, TypeError):
            lat, lon = None, None

        # Parse dates
        date_start = self._parse_year(item.get("date_start", item.get("date_from", item.get("min_date"))))
        date_end = self._parse_year(item.get("date_end", item.get("date_to", item.get("max_date"))))

        # Single date field
        if date_start is None and date_end is None:
            date_str = item.get("date", item.get("dating", ""))
            if date_str:
                dates = self._parse_date_range(str(date_str))
                date_start, date_end = dates

        return {
            "id": wreck_id,
            "name": item.get("name", item.get("wreck_name", item.get("title", ""))),
            "lat": lat,
            "lon": lon,
            "location_name": item.get("location", item.get("find_spot", item.get("region", ""))),
            "depth_meters": self._parse_float(item.get("depth", item.get("depth_m"))),
            "date_sunk_start": date_start,
            "date_sunk_end": date_end,
            "vessel_type": item.get("vessel_type", item.get("ship_type", "")),
            "cargo_type": item.get("cargo", item.get("cargo_type", "")),
            "cargo_origin": item.get("cargo_origin", ""),
            "cargo_destination": item.get("cargo_destination", item.get("destination", "")),
            "nationality": item.get("nationality", item.get("origin", "")),
            "construction": item.get("construction", item.get("hull_type", "")),
            "length_meters": self._parse_float(item.get("length", item.get("length_m"))),
            "preservation": item.get("preservation", item.get("condition", "")),
            "excavated": item.get("excavated", False),
            "bibliography": item.get("bibliography", item.get("references", "")),
            "notes": item.get("notes", item.get("description", "")),
        }

    def _parse_geojson_feature(self, feature: Dict) -> Optional[Dict]:
        """Parse a GeoJSON feature."""
        if not feature or feature.get("type") != "Feature":
            return None

        props = feature.get("properties", {})
        geom = feature.get("geometry", {})

        # Get coordinates
        lat, lon = None, None
        if geom.get("type") == "Point":
            coords = geom.get("coordinates", [])
            if len(coords) >= 2:
                lon, lat = coords[0], coords[1]

        props["lat"] = lat
        props["lon"] = lon

        return self._parse_wreck_item(props)

    def _parse_table_row(self, row: Dict) -> Optional[Dict]:
        """Parse a table row."""
        if not row:
            return None

        return self._parse_wreck_item(row)

    def _parse_year(self, value) -> Optional[int]:
        """Parse a year value."""
        if value is None:
            return None

        try:
            return int(float(value))
        except (ValueError, TypeError):
            return None

    def _parse_float(self, value) -> Optional[float]:
        """Parse a float value."""
        if value is None:
            return None

        try:
            return float(value)
        except (ValueError, TypeError):
            return None

    def _parse_date_range(self, date_str: str) -> tuple:
        """Parse a date range string like '200-100 BCE' or '1st c. CE'."""
        if not date_str:
            return None, None

        # Handle BCE/CE markers
        is_bce = "bce" in date_str.lower() or "bc" in date_str.lower()

        # Extract numbers
        numbers = re.findall(r'\d+', date_str)
        if not numbers:
            return None, None

        nums = [int(n) for n in numbers]

        # Apply BCE negative sign
        if is_bce:
            nums = [-n for n in nums]

        if len(nums) == 1:
            return nums[0], nums[0]
        elif len(nums) >= 2:
            return min(nums), max(nums)

        return None, None

    def parse(self, raw_data_path: Path) -> Iterator[ParsedSite]:
        """Parse OXREP shipwreck data into ParsedSite objects."""
        logger.info(f"Parsing OXREP shipwreck data from {raw_data_path}")

        with open(raw_data_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        wrecks = data.get("shipwrecks", [])
        logger.info(f"Processing {len(wrecks)} shipwrecks")

        for wreck in wrecks:
            if wreck.get("needs_manual_fetch"):
                continue

            site = self._wreck_to_site(wreck)
            if site:
                yield site

    def _wreck_to_site(self, wreck: Dict) -> Optional[ParsedSite]:
        """Convert a shipwreck to a ParsedSite."""
        lat = wreck.get("lat")
        lon = wreck.get("lon")

        if lat is None or lon is None:
            return None

        # Validate coordinates
        if not (-90 <= lat <= 90) or not (-180 <= lon <= 180):
            return None

        name = wreck.get("name") or f"Shipwreck at {wreck.get('location_name', 'Unknown')}"

        # Build description
        desc_parts = []
        if wreck.get("vessel_type"):
            desc_parts.append(f"Type: {wreck['vessel_type']}")
        if wreck.get("cargo_type"):
            desc_parts.append(f"Cargo: {wreck['cargo_type']}")
        if wreck.get("depth_meters"):
            desc_parts.append(f"Depth: {wreck['depth_meters']}m")
        if wreck.get("notes"):
            desc_parts.append(wreck["notes"][:200])

        return ParsedSite(
            source_id=wreck["id"],
            name=name,
            lat=lat,
            lon=lon,
            alternative_names=[],
            description="; ".join(desc_parts) if desc_parts else None,
            site_type="shipwreck",
            period_start=wreck.get("date_sunk_start"),
            period_end=wreck.get("date_sunk_end"),
            period_name=None,
            precision_meters=100,  # Underwater locations often imprecise
            precision_reason="underwater",
            source_url=f"https://oxrep.classics.ox.ac.uk/databases/shipwrecks_database/?id={wreck['id']}",
            raw_data=wreck,
        )


def ingest_oxrep_shipwrecks(session=None, skip_fetch: bool = False) -> dict:
    """Run OXREP Shipwrecks ingestion."""
    with OXREPShipwrecksIngester(session=session) as ingester:
        result = ingester.run(skip_fetch=skip_fetch)
        return {
            "source": result.source_id,
            "success": result.success,
            "saved": result.records_saved,
            "failed": result.records_failed,
        }
